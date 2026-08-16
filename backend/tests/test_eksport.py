"""Testy eksportu listy dokumentów do XLSX.

Uruchom: pytest backend/tests/test_eksport.py -v

Sprawdzamy nie „czy plik się zapisał", tylko rzeczy, dla których w ogóle robi się
eksport do arkusza: czy kolumny są w kolejności z rejestru, czy daty i liczby są
NATYWNYMI typami Excela (bez tego sortowanie i filtrowanie nie działa) i czy nic
z listy nie ginie po cichu.
"""
import io
from datetime import date, datetime

import pytest
from openpyxl import load_workbook


def jako_data(wartosc):
    """Excel nie zna „daty bez godziny" — openpyxl oddaje zapisaną datę jako
    `datetime`. Dla testu liczy się, że to TYP DATY (sortowalny), a nie napis."""
    assert isinstance(wartosc, (date, datetime)), f"oczekiwano daty, jest {type(wartosc)}"
    return wartosc.date() if isinstance(wartosc, datetime) else wartosc

from app.eksport import (
    etykieta_pola, naglowek_pobierania, nazwa_arkusza, nazwa_pliku, slug_z_pytania,
    zbuduj_xlsx,
)

SCHEMATY = {
    "zarzadzenie": {
        "slug": "zarzadzenie",
        "name": "Zarządzenie",
        # Kolejność jak w rejestrze — arkusz ma ją odwzorować co do pola.
        "fields": [
            {"name": "data", "type": "date"},
            {"name": "numer_dokumentu", "type": "string"},
            {"name": "tytul", "type": "string"},
            {"name": "osoba_podpisujaca", "type": "string"},
        ],
    },
    "aneks": {
        "slug": "aneks", "name": "Aneks",
        "fields": [{"name": "numer", "type": "number"}],
    },
}

ZARZADZENIE = {
    "filename": "20.pdf", "doc_type": "zarzadzenie",
    "doc_fields": {"data": "2023-10-16", "numer_dokumentu": "20/2023",
                   "tytul": "Regulamin Pracy", "osoba_podpisujaca": "Adrian Sikorski"},
}


def wczytaj(dokumenty, schematy=SCHEMATY):
    return load_workbook(io.BytesIO(zbuduj_xlsx(dokumenty, schematy)))


class TestUkladArkusza:
    def test_naglowki_w_kolejnosci_rejestru(self):
        ws = wczytaj([ZARZADZENIE])["Zarządzenie"]
        assert [k.value for k in ws[1]] == [
            "L.p.", "Data", "Numer dokumentu", "Tytul", "Osoba podpisujaca", "Plik"]

    def test_jeden_typ_to_jeden_arkusz(self):
        assert wczytaj([ZARZADZENIE, ZARZADZENIE]).sheetnames == ["Zarządzenie"]

    def test_kazdy_typ_dostaje_wlasny_arkusz(self):
        aneks = {"filename": "a.pdf", "doc_type": "aneks", "doc_fields": {"numer": "3"}}
        assert wczytaj([ZARZADZENIE, aneks]).sheetnames == ["Zarządzenie", "Aneks"]

    def test_typ_spoza_rejestru_nie_ginie(self):
        """Dokument bez rozpoznanego typu ma trafić do arkusza ogólnego, a nie
        wypaść z eksportu — inaczej lista w arkuszu byłaby krótsza niż na ekranie."""
        obcy = {"filename": "x.pdf", "doc_type": "cos-nowego", "doc_fields": {}}
        wb = wczytaj([ZARZADZENIE, obcy])
        assert "Pozostałe" in wb.sheetnames
        assert wb["Pozostałe"].max_row == 2                    # nagłówek + wiersz
        assert wb["Pozostałe"].cell(row=2, column=2).value == "x.pdf"

    def test_pusta_lista_nie_wywala(self):
        assert wczytaj([]).sheetnames == ["Dokumenty"]

    def test_naglowek_zamrozony_i_filtr_zalozony(self):
        ws = wczytaj([ZARZADZENIE])["Zarządzenie"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None


class TestWartosci:
    def test_data_jest_data_a_nie_tekstem(self):
        """Klucz do sortowania w Excelu. Tekst „2023-10-16" sortuje się jak napis."""
        ws = wczytaj([ZARZADZENIE])["Zarządzenie"]
        assert jako_data(ws.cell(row=2, column=2).value) == date(2023, 10, 16)

    def test_liczba_jest_liczba(self):
        aneks = {"filename": "a.pdf", "doc_type": "aneks", "doc_fields": {"numer": "3"}}
        assert wczytaj([aneks])["Aneks"].cell(row=2, column=2).value == 3

    @pytest.mark.parametrize("zapis,oczekiwana", [
        ("2023-10-16", date(2023, 10, 16)),
        ("16.10.2023", date(2023, 10, 16)),
        ("2023/10/16", date(2023, 10, 16)),
    ])
    def test_rozne_zapisy_daty(self, zapis, oczekiwana):
        d = {**ZARZADZENIE, "doc_fields": {**ZARZADZENIE["doc_fields"], "data": zapis}}
        assert jako_data(wczytaj([d])["Zarządzenie"].cell(row=2, column=2).value) == oczekiwana

    def test_nierozpoznana_data_zostaje_tekstem(self):
        """Lepiej pokazać, co było w dokumencie, niż wyciąć wartość."""
        d = {**ZARZADZENIE, "doc_fields": {**ZARZADZENIE["doc_fields"], "data": "II kwartał"}}
        assert wczytaj([d])["Zarządzenie"].cell(row=2, column=2).value == "II kwartał"

    def test_brak_wartosci_to_pusta_komorka(self):
        """Nie „—": pusty znak psułby filtrowanie i sortowanie w arkuszu."""
        d = {**ZARZADZENIE, "doc_fields": {"numer_dokumentu": "1/2009"}}
        assert wczytaj([d])["Zarządzenie"].cell(row=2, column=2).value is None

    def test_kolejnosc_wierszy_jak_na_ekranie(self):
        a = {**ZARZADZENIE, "doc_fields": {**ZARZADZENIE["doc_fields"], "numer_dokumentu": "1/2010"}}
        b = {**ZARZADZENIE, "doc_fields": {**ZARZADZENIE["doc_fields"], "numer_dokumentu": "13/2009"}}
        ws = wczytaj([a, b])["Zarządzenie"]
        assert [ws.cell(row=r, column=3).value for r in (2, 3)] == ["1/2010", "13/2009"]

    def test_numeracja_od_jedynki_w_kazdym_arkuszu(self):
        aneks = {"filename": "a.pdf", "doc_type": "aneks", "doc_fields": {"numer": "3"}}
        wb = wczytaj([ZARZADZENIE, ZARZADZENIE, aneks])
        assert [wb["Zarządzenie"].cell(row=r, column=1).value for r in (2, 3)] == [1, 2]
        assert wb["Aneks"].cell(row=2, column=1).value == 1


class TestNazw:
    @pytest.mark.parametrize("wejscie,oczekiwane", [
        ("osoba_podpisujaca", "Osoba podpisujaca"),
        ("data", "Data"),
        ("", "—"),
    ])
    def test_etykieta_pola(self, wejscie, oczekiwane):
        assert etykieta_pola(wejscie) == oczekiwane

    def test_nazwa_arkusza_bez_znakow_zakazanych(self):
        assert "/" not in nazwa_arkusza("Umowy 2024/2025", set())

    def test_nazwa_arkusza_przycieta_do_limitu(self):
        assert len(nazwa_arkusza("x" * 60, set())) <= 31

    def test_nazwa_arkusza_unikalna(self):
        zajete = set()
        assert nazwa_arkusza("Umowa", zajete) == "Umowa"
        assert nazwa_arkusza("Umowa", zajete) == "Umowa (2)"

    def test_nazwa_pliku_z_pytania(self):
        """Nazwa ma mowic, CO jest w srodku. Data pobrania tego nie mowi."""
        assert nazwa_pliku([ZARZADZENIE], SCHEMATY, "wypisz zarządzenia 2009") ==             "zarządzenia-2009.xlsx"

    @pytest.mark.parametrize("pytanie,oczekiwana", [
        ("wypisz zarządzenia 2009", "zarządzenia-2009"),
        ("Pokaż wszystkie umowy z 2023", "umowy-z-2023"),
        ("zarzadzenia 2009", "zarzadzenia-2009"),          # bez ogonkow, jak wpisano
        ("wymień instrukcje", "instrukcje"),
        ("", ""),
        ("wypisz", ""),                                     # samo polecenie = brak nazwy
    ])
    def test_slug_z_pytania(self, pytanie, oczekiwana):
        assert slug_z_pytania(pytanie) == oczekiwana

    def test_dlugie_pytanie_przyciete_na_granicy_slowa(self):
        slug = slug_z_pytania("wypisz " + " ".join(["zarzadzenie"] * 12))
        assert len(slug) <= 60 and not slug.endswith("-") and "zarzadzenie" in slug
        assert slug.split("-")[-1] == "zarzadzenie"          # nie ucieta w polowie

    def test_bez_pytania_nazwa_po_typie(self):
        assert nazwa_pliku([ZARZADZENIE], SCHEMATY) == "zarządzenie.xlsx"

    def test_bez_pytania_i_przy_typach_mieszanych_nazwa_ogolna(self):
        aneks = {"filename": "a.pdf", "doc_type": "aneks", "doc_fields": {}}
        assert nazwa_pliku([ZARZADZENIE, aneks], SCHEMATY) == "lista-dokumentow.xlsx"


class TestNaglowkaPobierania:
    """Regresja z 2026-08-10: pierwszy eksport zwrocil 500. Arkusz budowal sie dobrze,
    ale nazwa „zarządzenie-….xlsx" nie przechodzila przez naglowek HTTP — te sa kodowane
    w latin-1 i „ą" wywracalo cala odpowiedz. Test pilnuje warunku, ktorego wczesniej
    nie sprawdzalem: czy naglowek DA SIE WYSLAC."""

    @pytest.mark.parametrize("nazwa", [
        "zarządzenie-2026-08-10.xlsx",
        "lista-dokumentow-2026-08-10.xlsx",
        "źdźbło-ćma-ĄĘŁŃÓŚŻ.xlsx",
    ])
    def test_naglowek_da_sie_zakodowac_w_latin1(self, nazwa):
        naglowek_pobierania(nazwa).encode("latin-1")      # rzuci, gdy wroci regresja

    def test_zawiera_oba_warianty_nazwy(self):
        n = naglowek_pobierania("zarządzenie-2026.xlsx")
        assert 'filename="zarzadzenie-2026.xlsx"' in n     # awaryjny, transliterowany
        assert "filename*=UTF-8''" in n                    # pelna nazwa dla przegladarki

    def test_nazwa_bez_ani_jednego_znaku_ascii(self):
        """Sama transliteracja moglaby dac pusty napis — wtedy potrzebny jest zapas."""
        assert 'filename="lista-dokumentow.xlsx"' in naglowek_pobierania("日本語.xlsx")


SCHEMAT_FAKTURY = {
    "faktura": {
        "slug": "faktura", "name": "Faktura",
        "fields": [{"name": "kwota_brutto", "type": "money"}],
    },
}


def faktura(zapis):
    return {"filename": "f.pdf", "doc_type": "faktura", "doc_fields": {"kwota_brutto": zapis}}


class TestKwoty:
    """Typ „money" istnieje po to, żeby kolumna z kwotami dała się ZSUMOWAĆ.
    Gdyby wartość trafiała do arkusza jako tekst, cały ten typ byłby ozdobą."""

    @pytest.mark.parametrize("zapis,oczekiwana", [
        ("1234,56", 1234.56),
        ("1 234,56 zł", 1234.56),          # spacja grupująca i symbol waluty
        ("1\u00a0234,56 PLN", 1234.56),    # spacja niełamliwa z dokumentu
        ("1.234,56", 1234.56),             # kropka jako tysiące, przecinek jako grosze
        ("1,234.56", 1234.56),             # zapis angielski — ostatni separator rządzi
        ("12.345.678", 12345678.0),        # same tysiące, bez groszy
        ("1.5", 1.5),                      # kropka dziesiętna, nie tysiące
        ("980", 980.0),
        ("-45,20", -45.2),
    ])
    def test_rozne_zapisy_kwoty(self, zapis, oczekiwana):
        ws = wczytaj([faktura(zapis)], SCHEMAT_FAKTURY)["Faktura"]
        assert ws.cell(row=2, column=2).value == pytest.approx(oczekiwana)

    def test_kwota_dostaje_format_z_groszami(self):
        ws = wczytaj([faktura("1234,5")], SCHEMAT_FAKTURY)["Faktura"]
        assert ws.cell(row=2, column=2).number_format == "#,##0.00"

    def test_nierozpoznana_kwota_zostaje_tekstem(self):
        # Lepiej pokazać, co było w dokumencie, niż zamienić śmieć na zero.
        ws = wczytaj([faktura("do uzgodnienia")], SCHEMAT_FAKTURY)["Faktura"]
        komorka = ws.cell(row=2, column=2)
        assert komorka.value == "do uzgodnienia"
        assert komorka.number_format != "#,##0.00", "tekst nie może udawać kwoty"

    def test_brak_kwoty_to_pusta_komorka(self):
        ws = wczytaj([faktura(None)], SCHEMAT_FAKTURY)["Faktura"]
        assert ws.cell(row=2, column=2).value is None
