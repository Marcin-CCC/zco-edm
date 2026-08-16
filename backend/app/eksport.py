"""Eksport listy dokumentów do arkusza XLSX.

Po co: odpowiedź czatu typu LISTA („wypisz zarządzenia 2009") pokazuje dokumenty
na ekranie, ale pracownik i tak potrzebuje ich w Excelu — do rejestru, do przesłania
dalej, do przefiltrowania po dacie. Kopiowanie z przeglądarki gubi strukturę.

Kolumny biorą się z REJESTRU SCHEMATÓW (`doc_type_schemas.fields`), w kolejności
tam zapisanej. To celowo jedno źródło prawdy: kolejność pól widoczna w podglądzie
dokumentu jest tą samą, którą zobaczy się w arkuszu. Zmiana układu kolumn = zmiana
kolejności pól w Administracji, a nie osobny kreator eksportu.

Jeden arkusz na TYP dokumentu. Lista mieszana („dokumenty z 2009") nie ma wspólnego
zestawu pól, więc rozdzielamy ją na arkusze; lista jednorodna — najczęstszy przypadek
— daje po prostu jeden arkusz.

Wartości zapisujemy w NATYWNYCH typach Excela: data jako data, liczba jako liczba.
Bez tego sortowanie i filtrowanie w arkuszu nie działa — a to główny powód, dla
którego ktoś woli xlsx od przepisywania z ekranu.
"""

import io
import re
from datetime import datetime
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Excel nie przyjmuje tych znaków w nazwie arkusza ani nazw dłuższych niż 31 znaków.
_ZAKAZANE_W_NAZWIE = re.compile(r"[\[\]:*?/\\]")
MAX_NAZWA_ARKUSZA = 31

NAGLOWEK_TLO = PatternFill("solid", fgColor="F3F4F6")
NAGLOWEK_FONT = Font(bold=True)

# Format kolumn kwotowych. Bez symbolu waluty: pole „Kwota" bywa w euro tak samo
# jak w złotych, a dopisany na sztywno symbol byłby wtedy po prostu nieprawdą.
# Kod formatu jest niezależny od języka — Excel pokazuje separatory wg ustawień
# użytkownika, więc w polskiej wersji wyjdzie „1 234,56".
FORMAT_KWOTY = "#,##0.00"


def nazwa_arkusza(nazwa: str, zajete: set[str]) -> str:
    """Nazwa arkusza dopuszczalna dla Excela i niepowtarzalna w skoroszycie."""
    czysta = _ZAKAZANE_W_NAZWIE.sub("-", (nazwa or "Dokumenty").strip()) or "Dokumenty"
    czysta = czysta[:MAX_NAZWA_ARKUSZA]
    if czysta not in zajete:
        zajete.add(czysta)
        return czysta
    # Kolizja: dokładamy numer, przycinając nazwę tak, by zmieścić się w limicie.
    for i in range(2, 100):
        sufiks = f" ({i})"
        kandydat = czysta[: MAX_NAZWA_ARKUSZA - len(sufiks)] + sufiks
        if kandydat not in zajete:
            zajete.add(kandydat)
            return kandydat
    zajete.add(czysta)
    return czysta


def etykieta_pola(nazwa: str) -> str:
    """`osoba_podpisujaca` → `Osoba podpisujaca`.

    Rejestr trzyma nazwy techniczne (bez spacji i ogonków), bo służą też za klucze
    w danych. Do nagłówka arkusza zamieniamy je na czytelne; gdyby kiedyś doszły
    do rejestru osobne etykiety, to jedyne miejsce do zmiany.
    """
    return (nazwa or "").replace("_", " ").strip().capitalize() or "—"


def _wartosc(surowa, typ: str):
    """Wartość w typie natywnym Excela, gdy tylko da się ją rozpoznać."""
    if surowa is None or surowa == "":
        return None
    tekst = str(surowa).strip()
    if typ == "date":
        for wzor in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(tekst, wzor).date()
            except ValueError:
                continue
        return tekst                      # nierozpoznany zapis — lepiej tekst niż nic
    if typ == "number":
        try:
            liczba = float(tekst.replace(",", "."))
            return int(liczba) if liczba.is_integer() else liczba
        except ValueError:
            return tekst
    if typ == "money":
        kwota = _kwota(tekst)
        return tekst if kwota is None else kwota
    return tekst


# Symbole i skróty walut, które model wyciąga razem z liczbą („1 234,56 zł").
_WALUTY = re.compile(r"(?i)\b(zl|zł|pln|eur|usd|gbp|chf)\b|[€$£]")


def _kwota(tekst: str) -> float | None:
    """„1 234,56 zł" → 1234.56. ``None``, gdy to nie jest kwota.

    Kwota musi trafić do arkusza jako LICZBA, nie tekst — inaczej kolumna się nie
    zsumuje, a to jedyny powód, dla którego „Kwota" jest osobnym typem obok
    „Liczby". Rozdzielamy separatory po polsku: przecinek jest dziesiętny, spacja
    (także niełamliwa) grupuje tysiące.
    """
    czysty = _WALUTY.sub("", tekst)
    czysty = re.sub(r"[\s\u00a0\u202f]", "", czysty).strip()
    if not czysty:
        return None

    if "," in czysty and "." in czysty:
        # Separatorem dziesiętnym jest ten ostatni; drugi grupuje tysiące.
        if czysty.rfind(",") > czysty.rfind("."):
            czysty = czysty.replace(".", "").replace(",", ".")
        else:
            czysty = czysty.replace(",", "")
    elif "," in czysty:
        czysty = czysty.replace(",", ".")
    elif re.fullmatch(r"-?\d{1,3}(\.\d{3})+", czysty):
        # „1.234" i „12.345.678" to po polsku tysiące, nie ułamek. Trzy cyfry po
        # kropce i cyfry przed nią — inaczej „1.5" byłoby czytane jako 15 setek.
        czysty = czysty.replace(".", "")

    try:
        return float(czysty)
    except ValueError:
        return None


def _typ_pola(definicja: dict) -> str:
    """`enum:a,b` traktujemy jak tekst — liczy się tylko część przed dwukropkiem."""
    return (definicja.get("type") or "string").split(":", 1)[0]


def zbuduj_xlsx(dokumenty: list[dict], schematy: dict[str, dict]) -> bytes:
    """Zbuduj skoroszyt z listy dokumentów.

    `dokumenty` — [{"filename", "doc_type", "doc_fields"}] W KOLEJNOŚCI WYŚWIETLENIA.
    Kolejność ma znaczenie: użytkownik widział listę posortowaną w konkretny sposób
    i tego samego oczekuje w arkuszu.

    `schematy` — {slug: {"name": str, "fields": [{"name","type"}, ...]}}.
    Typ spoza rejestru (albo dokument bez typu) trafia do arkusza „Pozostałe"
    z kolumnami ogólnymi — nie gubimy takich pozycji po cichu.
    """
    # Grupowanie z zachowaniem kolejności pierwszego wystąpienia typu.
    grupy: dict[str, list[dict]] = {}
    for d in dokumenty:
        slug = d.get("doc_type") or ""
        klucz = slug if slug in schematy else ""
        grupy.setdefault(klucz, []).append(d)

    wb = Workbook()
    wb.remove(wb.active)                  # domyślny pusty arkusz
    zajete: set[str] = set()

    for slug, pozycje in grupy.items():
        schemat = schematy.get(slug)
        pola = list(schemat["fields"]) if schemat else []
        tytul = schemat["name"] if schemat else "Pozostałe"
        ws = wb.create_sheet(nazwa_arkusza(tytul, zajete))

        naglowki = ["L.p."] + [etykieta_pola(p.get("name")) for p in pola] + ["Plik"]
        ws.append(naglowki)
        for komorka in ws[1]:
            komorka.font = NAGLOWEK_FONT
            komorka.fill = NAGLOWEK_TLO
            komorka.alignment = Alignment(vertical="center")

        # Kolumny kwotowe dostają format z groszami i separatorem tysięcy — bez
        # niego arkusz pokazuje „1234.5" tam, gdzie w dokumencie było „1 234,50".
        kolumny_kwot = [i for i, p in enumerate(pola, start=2) if _typ_pola(p) == "money"]

        for nr, d in enumerate(pozycje, 1):
            wartosci = d.get("doc_fields") or {}
            wiersz = [nr]
            for p in pola:
                wiersz.append(_wartosc(wartosci.get(p.get("name")), _typ_pola(p)))
            wiersz.append(d.get("filename") or "")
            ws.append(wiersz)
            for kol in kolumny_kwot:
                komorka = ws.cell(row=ws.max_row, column=kol)
                # Format tylko dla wartości liczbowych: nierozpoznany zapis został
                # tekstem i format walutowy zrobiłby z niego zero.
                if isinstance(komorka.value, (int, float)):
                    komorka.number_format = FORMAT_KWOTY

        # Filtr i zamrożony nagłówek — arkusz ma służyć do przeglądania rejestru,
        # a nie tylko do archiwizacji.
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(naglowki))}{ws.max_row}"

        # Szerokości kolumn z zawartości; górny limit, bo tytuły dokumentów bywają
        # bardzo długie i rozjechałyby arkusz.
        for i, _ in enumerate(naglowki, 1):
            litera = get_column_letter(i)
            najdluzsza = max(
                (len(str(k.value)) for k in ws[litera] if k.value is not None),
                default=8,
            )
            ws.column_dimensions[litera].width = min(max(najdluzsza + 2, 8), 55)
        ws.column_dimensions["A"].width = 6      # L.p. nie potrzebuje szerokości

    if not wb.sheetnames:                 # pusta lista — oddajemy czytelny, pusty arkusz
        ws = wb.create_sheet("Dokumenty")
        ws.append(["L.p.", "Plik"])

    bufor = io.BytesIO()
    wb.save(bufor)
    return bufor.getvalue()


# Polskie znaki na ASCII — do awaryjnej nazwy pliku w nagłówku HTTP.
_TRANSLITERACJA = str.maketrans("ąćęłńóśźżĄĆĘŁŃÓŚŹŻ", "acelnoszzACELNOSZZ")


def naglowek_pobierania(nazwa: str) -> str:
    """Wartość nagłówka `Content-Disposition` dla nazwy z polskimi znakami.

    Nagłówki HTTP kodowane są w latin-1, więc nazwa z „ą" przewraca CAŁĄ odpowiedź
    (`UnicodeEncodeError`). Zmierzone boleśnie: pierwszy eksport zwrócił 500, choć sam
    arkusz budował się poprawnie — błąd wyszedł dopiero przy wysyłaniu nagłówka.

    Dajemy dwa warianty naraz, zgodnie z RFC 6266: `filename` z transliteracją jako
    awaryjny i `filename*` w UTF-8 dla przeglądarek, które umieją pokazać ogonki.
    """
    awaryjna = nazwa.translate(_TRANSLITERACJA).encode("ascii", "ignore").decode()
    # Sprawdzamy TRZON, nie całość: dla nazwy bez znaków ASCII zostaje samo „.xlsx",
    # co jest niepustym napisem, ale bezużyteczną nazwą pliku.
    trzon = awaryjna.rsplit(".", 1)[0] if "." in awaryjna else awaryjna
    if not trzon.strip(" -_"):
        awaryjna = "lista-dokumentow.xlsx"
    return f"attachment; filename=\"{awaryjna}\"; filename*=UTF-8''{quote(nazwa)}"


# Polecenia otwierające pytanie („wypisz zarządzenia 2009"). W nazwie pliku są
# szumem — opisują, o co użytkownik prosił, a nie CO dostał.
_POLECENIA = {
    "wypisz", "pokaz", "wymien", "znajdz", "wyszukaj", "wylistuj", "podaj", "daj",
    "otworz", "przeslij", "eksportuj", "lista", "liste", "wszystkie", "prosze",
}
_SLOWO_NAZWY = re.compile(r"[0-9A-Za-zĄĆĘŁŃÓŚŹŻąćęłńóśźż]+")
MAX_DLUGOSC_NAZWY = 60


def slug_z_pytania(pytanie: str | None) -> str:
    """„wypisz zarządzenia 2009" → „zarządzenia-2009".

    Nazwa pliku ma mówić, CO jest w środku. Data pobrania tego nie mówi — mówi tylko,
    kiedy ktoś kliknął, a to widać w systemie plików.

    Polecenia z początku odcinamy, bo w nazwie są szumem. Nie tłumaczymy ani nie
    poprawiamy pisowni: nazwa ma odwzorowywać pytanie, więc kto napisał „zarzadzenia",
    dostanie „zarzadzenia".
    """
    slowa = _SLOWO_NAZWY.findall(pytanie or "")
    while slowa and slowa[0].lower().translate(_TRANSLITERACJA) in _POLECENIA:
        slowa.pop(0)
    slug = "-".join(slowa).lower()
    if len(slug) > MAX_DLUGOSC_NAZWY:
        # Tniemy na granicy słowa, żeby nazwa nie kończyła się w połowie wyrazu.
        slug = slug[:MAX_DLUGOSC_NAZWY].rsplit("-", 1)[0]
    return slug.strip("-")


def nazwa_pliku(dokumenty: list[dict], schematy: dict[str, dict],
                pytanie: str | None = None) -> str:
    """Nazwa pobieranego pliku: z treści pytania, a gdy się nie da — z typu."""
    slug = slug_z_pytania(pytanie)
    if slug:
        return f"{slug}.xlsx"

    typy = {d.get("doc_type") for d in dokumenty}
    if len(typy) == 1:
        schemat = schematy.get(typy.pop() or "")
        if schemat:
            czysty = re.sub(r"[^\w-]+", "-", schemat["name"].lower(), flags=re.UNICODE)
            return f"{czysty.strip('-')}.xlsx"
    return "lista-dokumentow.xlsx"
